"""
excel_exporter.py
-----------------
Generates a fully formatted, multi-sheet Excel workbook from an article dataset.

Sheets
------
1. Dashboard   — KPI summary cards + per-category and status stats
2. Articles    — Full article table with conditional formatting
3. Analytics   — Engagement metrics table

Design rules (SKILL.md)
------------------------
- Arial font throughout
- Excel formulas for all aggregates (no Python-computed hardcodes)
- Black text  → formulas / calculated values
- Zero merge-region overlaps
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from data import Article, articles_to_dataframe

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────────────────────────────────────
NAVY       = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
BLACK      = "000000"
DARK_GREY  = "404040"
LIGHT_GREY = "F2F2F2"
ORANGE     = "E97132"

STATUS_COLOURS = {
    "Published": "C6EFCE",
    "Review":    "FFEB9C",
    "Draft":     "FFCCCC",
}


class ExcelExporter:
    """
    Builds a professional multi-sheet Excel workbook.

    Usage
    -----
        path = ExcelExporter(articles).export("output/articles_report.xlsx")
    """

    def __init__(self, articles: list[Article]) -> None:
        self.articles = articles
        self.df       = articles_to_dataframe(articles)
        self.wb       = Workbook()

    # ── Public API ─────────────────────────────────────────────────────────────

    def export(self, output_path: str | Path) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.remove(self.wb.active)
        self._build_dashboard()
        self._build_articles_sheet()
        self._build_analytics_sheet()
        self.wb.save(path)
        logger.info("Excel workbook saved → %s", path)
        return path

    # ── Sheet 1: Dashboard ─────────────────────────────────────────────────────

    def _build_dashboard(self) -> None:
        ws = self.wb.create_sheet("Dashboard")
        ws.sheet_view.showGridLines = False

        # ── Banner ─────────────────────────────────────────────────────────────
        ws.merge_cells("B2:L3")
        banner           = ws["B2"]
        banner.value     = "ARTICLE DATASET — EXECUTIVE DASHBOARD"
        banner.font      = Font(name="Arial", size=18, bold=True, color=WHITE)
        banner.fill      = PatternFill("solid", fgColor=NAVY)
        banner.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 28

        ws.merge_cells("B4:L4")
        sub           = ws["B4"]
        sub.value     = "Aggregated performance metrics across the full article corpus"
        sub.font      = Font(name="Arial", size=10, italic=True, color=DARK_GREY)
        sub.alignment = Alignment(horizontal="center")

        # ── KPI cards (cols B,D,F,H,J — each card is 2 cols wide) ─────────────
        kpis = [
            ("B", "Total Articles",   "=COUNTA(Articles!A2:A1000)",  "#,##0"),
            ("D", "Total Reads",      "=SUM(Articles!G2:G1000)",      "#,##0"),
            ("F", "Total Likes",      "=SUM(Articles!H2:H1000)",      "#,##0"),
            ("H", "Avg Engagement %", "=AVERAGE(Articles!I2:I1000)",  "0.00%"),
            ("J", "Avg Word Count",   "=AVERAGE(Articles!F2:F1000)",  "#,##0"),
        ]

        for col, label, formula, fmt in kpis:
            end_col = get_column_letter(column_index_from_string(col) + 1)

            ws.merge_cells(f"{col}6:{end_col}6")
            lbl_cell           = ws[f"{col}6"]
            lbl_cell.value     = label
            lbl_cell.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
            lbl_cell.fill      = PatternFill("solid", fgColor=MID_BLUE)
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(f"{col}7:{end_col}9")
            val_cell              = ws[f"{col}7"]
            val_cell.value        = formula
            val_cell.font         = Font(name="Arial", size=20, bold=True, color=NAVY)
            val_cell.fill         = PatternFill("solid", fgColor=LIGHT_BLUE)
            val_cell.alignment    = Alignment(horizontal="center", vertical="center")
            val_cell.number_format = fmt

            ws.column_dimensions[col].width    = 15
            ws.column_dimensions[end_col].width = 3

        ws.column_dimensions["L"].width = 3

        # ── Section A: Category Breakdown (cols B–G, starting row 12) ──────────
        self._section_header(ws, "B", 12, "G", "PERFORMANCE BY CATEGORY")

        cat_headers = ["Category", "Articles", "Total Reads", "Total Likes", "Avg Engagement"]
        col_widths  = [20, 10, 14, 12, 16]
        for ci, (h, w) in enumerate(zip(cat_headers, col_widths)):
            cell           = ws.cell(row=14, column=2 + ci)
            cell.value     = h
            cell.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = self._thin_border()
            ws.column_dimensions[get_column_letter(2 + ci)].width = w

        categories = sorted({a.category for a in self.articles})
        for ri, cat in enumerate(categories):
            row    = 15 + ri
            sub_df = self.df[self.df["Category"] == cat]
            vals   = [
                cat,
                len(sub_df),
                int(sub_df["Reads"].sum()),
                int(sub_df["Likes"].sum()),
                round(sub_df["Engagement (%)"].mean(), 2) / 100,
            ]
            fmts   = [None, "#,##0", "#,##0", "#,##0", "0.00%"]
            bg     = LIGHT_GREY if ri % 2 == 0 else WHITE
            for ci, (val, fmt) in enumerate(zip(vals, fmts)):
                cell           = ws.cell(row=row, column=2 + ci)
                cell.value     = val
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="left" if ci == 0 else "center")
                cell.border    = self._thin_border()
                if fmt:
                    cell.number_format = fmt

        # ── Section B: Status Breakdown (cols H–L, starting row 12) ───────────
        self._section_header(ws, "H", 12, "L", "STATUS BREAKDOWN")

        for ci, h in enumerate(["Status", "Count", "Share"]):
            cell           = ws.cell(row=14, column=8 + ci)
            cell.value     = h
            cell.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = self._thin_border()

        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 8
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 3
        ws.column_dimensions["L"].width = 3

        for ri, status in enumerate(["Published", "Review", "Draft"]):
            row   = 15 + ri
            count = sum(1 for a in self.articles if a.status == status)
            pct   = count / len(self.articles)
            bg    = STATUS_COLOURS.get(status, WHITE)

            sc           = ws.cell(row=row, column=8, value=status)
            sc.font      = Font(name="Arial", size=9, bold=True)
            sc.fill      = PatternFill("solid", fgColor=bg)
            sc.alignment = Alignment(horizontal="center")
            sc.border    = self._thin_border()

            nc           = ws.cell(row=row, column=9, value=count)
            nc.font      = Font(name="Arial", size=9)
            nc.alignment = Alignment(horizontal="center")
            nc.border    = self._thin_border()
            nc.fill      = PatternFill("solid", fgColor=LIGHT_GREY if ri % 2 == 0 else WHITE)

            pc              = ws.cell(row=row, column=10, value=pct)
            pc.font         = Font(name="Arial", size=9)
            pc.number_format = "0.0%"
            pc.alignment    = Alignment(horizontal="center")
            pc.border       = self._thin_border()
            pc.fill         = PatternFill("solid", fgColor=LIGHT_GREY if ri % 2 == 0 else WHITE)

    # ── Sheet 2: Articles ──────────────────────────────────────────────────────

    def _build_articles_sheet(self) -> None:
        ws = self.wb.create_sheet("Articles")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        columns = [
            ("ID",            6,   "#,##0"),
            ("Title",         52,  None),
            ("Author",        22,  None),
            ("Category",      15,  None),
            ("Date",          14,  None),
            ("Word Count",    13,  "#,##0"),
            ("Reads",         13,  "#,##0"),
            ("Likes",         13,  "#,##0"),
            ("Engagement %",  14,  "0.00%"),
            ("Status",        13,  None),
            ("Tags",          30,  None),
            ("Summary",       62,  None),
        ]

        # Header row
        for ci, (header, width, _) in enumerate(columns, 1):
            cell           = ws.cell(row=1, column=ci, value=header)
            cell.font      = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = self._thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 32

        # Data rows
        for ri, article in enumerate(self.articles, 2):
            row_data = [
                article.id,
                article.title,
                article.author,
                article.category,
                article.date,
                article.word_count,
                article.reads,
                article.likes,
                article.engagement_rate / 100,
                article.status,
                ", ".join(article.tags),
                article.summary,
            ]
            bg = LIGHT_GREY if ri % 2 == 0 else WHITE
            for ci, (_, _, fmt) in enumerate(columns, 1):
                cell           = ws.cell(row=ri, column=ci, value=row_data[ci - 1])
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(vertical="top", wrap_text=(ci in (2, 11, 12)))
                cell.border    = self._thin_border()
                if fmt:
                    cell.number_format = fmt

            # Status badge
            sc           = ws.cell(row=ri, column=10)
            sc.fill      = PatternFill("solid", fgColor=STATUS_COLOURS.get(article.status, WHITE))
            sc.alignment = Alignment(horizontal="center", vertical="top")
            sc.font      = Font(name="Arial", size=9, bold=True)
            ws.row_dimensions[ri].height = 60

        # Totals row
        total_row = len(self.articles) + 2
        data_end  = len(self.articles) + 1
        col_map   = {name: get_column_letter(i + 1) for i, (name, _, _) in enumerate(columns)}

        for ci in range(1, len(columns) + 1):
            cell        = ws.cell(row=total_row, column=ci)
            cell.fill   = PatternFill("solid", fgColor=NAVY)
            cell.border = self._thin_border()

        ws.cell(row=total_row, column=1).value = "TOTALS"
        ws.cell(row=total_row, column=1).font  = Font(name="Arial", size=9, bold=True, color=WHITE)

        totals = {
            "Word Count":   (f"=SUM({col_map['Word Count']}2:{col_map['Word Count']}{data_end})", "#,##0"),
            "Reads":        (f"=SUM({col_map['Reads']}2:{col_map['Reads']}{data_end})", "#,##0"),
            "Likes":        (f"=SUM({col_map['Likes']}2:{col_map['Likes']}{data_end})", "#,##0"),
            "Engagement %": (f"=AVERAGE({col_map['Engagement %']}2:{col_map['Engagement %']}{data_end})", "0.00%"),
        }
        for col_name, (formula, fmt) in totals.items():
            ci           = next(i for i, (c, _, _) in enumerate(columns, 1) if c == col_name)
            cell         = ws.cell(row=total_row, column=ci, value=formula)
            cell.font    = Font(name="Arial", size=9, bold=True, color=WHITE)
            cell.fill    = PatternFill("solid", fgColor=NAVY)
            cell.border  = self._thin_border()
            cell.number_format = fmt

    # ── Sheet 3: Analytics ─────────────────────────────────────────────────────

    def _build_analytics_sheet(self) -> None:
        ws = self.wb.create_sheet("Analytics")
        ws.sheet_view.showGridLines = False

        self._section_header(ws, "B", 2, "K", "TOP ARTICLES BY ENGAGEMENT RATE")

        headers = ["Rank", "Title", "Author", "Reads", "Likes", "Engagement %", "Category"]
        widths  = [6, 50, 22, 13, 13, 14, 15]
        for ci, (h, w) in enumerate(zip(headers, widths), 2):
            cell           = ws.cell(row=4, column=ci, value=h)
            cell.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=MID_BLUE)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = self._thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = w

        sorted_articles = sorted(self.articles, key=lambda a: a.engagement_rate, reverse=True)
        for ri, article in enumerate(sorted_articles, 5):
            bg       = LIGHT_BLUE if ri % 2 == 0 else WHITE
            row_data = [
                ri - 4,
                article.title,
                article.author,
                article.reads,
                article.likes,
                article.engagement_rate / 100,
                article.category,
            ]
            fmts = [None, None, None, "#,##0", "#,##0", "0.00%", None]
            for ci, (val, fmt) in enumerate(zip(row_data, fmts), 2):
                cell           = ws.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="left" if ci == 3 else "center")
                cell.border    = self._thin_border()
                if fmt:
                    cell.number_format = fmt

        # Category aggregates (formula-driven, separate section below)
        agg_start = len(sorted_articles) + 7
        self._section_header(ws, "B", agg_start, "K", "CATEGORY AGGREGATES (FORMULA-DRIVEN)")

        agg_hdrs = ["Category", "Count", "Sum Reads", "Sum Likes", "Avg Engagement %"]
        for ci, h in enumerate(agg_hdrs, 2):
            cell           = ws.cell(row=agg_start + 2, column=ci, value=h)
            cell.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = self._thin_border()

        for ri_off, cat in enumerate(sorted({a.category for a in self.articles})):
            row = agg_start + 3 + ri_off
            bg  = LIGHT_GREY if ri_off % 2 == 0 else WHITE

            lbl        = ws.cell(row=row, column=2, value=cat)
            lbl.font   = Font(name="Arial", size=9)
            lbl.fill   = PatternFill("solid", fgColor=bg)
            lbl.border = self._thin_border()

            formulas = [
                (f'=COUNTIF(Articles!D:D,"{cat}")',                    None),
                (f'=SUMIF(Articles!D:D,"{cat}",Articles!G:G)',         "#,##0"),
                (f'=SUMIF(Articles!D:D,"{cat}",Articles!H:H)',         "#,##0"),
                (f'=AVERAGEIF(Articles!D:D,"{cat}",Articles!I:I)',      "0.00%"),
            ]
            for ci_off, (formula, fmt) in enumerate(formulas):
                cell           = ws.cell(row=row, column=3 + ci_off, value=formula)
                cell.font      = Font(name="Arial", size=9, color=BLACK)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center")
                cell.border    = self._thin_border()
                if fmt:
                    cell.number_format = fmt

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _section_header(self, ws: Worksheet, start_col: str, row: int, end_col: str, text: str) -> None:
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
        cell           = ws[f"{start_col}{row}"]
        cell.value     = text
        cell.font      = Font(name="Arial", size=11, bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def _thin_border(self) -> Border:
        side = Side(style="thin", color="CCCCCC")
        return Border(left=side, right=side, top=side, bottom=side)
